#Search restaurants list with Postcodes

from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import os
# import xlsxwriter
import openpyxl
import random


class scrapeDetails():
    def Details(parcels):
        headerList = ['ALT KEY', 'SITE ADDRESS', 'LEGAL DESCRIPTION', 'LAND USE', 'OWNER 1', 'OWNER1 ADDRESS',
                    'OWNER 2', 'OWNER2 ADDRESS','OWNER3', 'OWNER3 ADDRESS', 'OCCUPANCY TYPE', 'YEAR BUILT',
                    'SALE DATE', 'SALE PRICE', 'LAND VALUE', 'BLDG VALUE', 'JUST VALUE']
        
        
        currentPath = os.path.dirname(os.path.realpath(__file__)) + '/chromedriver.exe'
        driver = webdriver.Chrome(executable_path= currentPath)
        # workbook = xlsxwriter.Workbook('Parcels.xlsx') 
        # worksheet = workbook.add_worksheet("Details")
        url = 'https://beacon.schneidercorp.com/Application.aspx?AppID=1024&PageType=Search'
        workbook = openpyxl.load_workbook(filename = 'SCOTT COUNTY IA.xlsx')
        worksheet = workbook['Deduped  Combined Sheets']
        # print(worksheet.max_row)
        # return
        row_count = 1
        #Set Sheet Colum Width
        
        # worksheet.column_dimensions['AX'] = 10
       
        # worksheet.column_dimensions['BJ'] = 12
        #Set Sheet Colum Header
        # worksheet.cell(row = row_count, column = 49).value =  'ALT KEY'
        # worksheet.cell(row = row_count, column = 50).value = "SITE ADDRESS"
        # worksheet.cell(row = row_count, column = 51).value = "LEGAL DESCRIPTION"
        # worksheet.cell(row = row_count, column = 52).value = 'LAND USE'
        # worksheet.cell(row = row_count, column = 53).value = "OWNER"
        # worksheet.cell(row = row_count, column = 54).value = "OWNER ADDRESS"
        # worksheet.cell(row = row_count, column = 55).value = 'OCCUPANCY TYPE'
        # worksheet.cell(row = row_count, column = 56).value = "YEAR BUILT"
        # worksheet.cell(row = row_count, column = 57).value = "SALE DATE"
        # worksheet.cell(row = row_count, column = 58).value = 'SALE PRICE'
        # worksheet.cell(row = row_count, column = 59).value = "LAND VALUE"
        # worksheet.cell(row = row_count, column = 60).value = "BLDG VALUE"
        # worksheet.cell(row = row_count, column = 61).value = "JUST VALUE"
        for idx in range(0, len(headerList)):
            worksheet.cell(row = row_count, column = idx + 49).value = headerList[idx]
        # Setting Random Actions.
        collapseID1 = 'ctlBodyPane_ctl'
        collapseID2 = '_btnToggleVis'
        i = 0
        for row_count in range(2, worksheet.max_row):
            # if i > 50:
                # break
            i += 1
            print("Scraped: {} items".format(i))
            # Initialize cell value lists
            valueList = []
            for idx in range(0,18):
                valueList.append(' ')
            randomTime = random.randint(1,3) * 2
            randomCnum = random.randint(1,25)
            randomoption = random.randint(1,5)
            # row_count += 1
            time.sleep(randomTime)
            driver.get(url)
            parcel = worksheet.cell(row = row_count, column = 1).value
            # Check Agreement Button
            if len(driver.find_elements_by_xpath("//div[@class = 'modal-dialog']")):
                agreeButton = driver.find_element_by_xpath("//div[@class = 'modal-dialog']/div[@class = 'modal-content']/div[@class = 'modal-focus-target']/div[@class = 'modal-footer']/a[@class = 'btn btn-primary button-1']")
                agreeButton.click()
            #Search with Parcel
            search_box = driver.find_element_by_xpath("//input[@id='ctlBodyPane_ctl04_ctl01_txtParcelID']")
            search_box.clear()
            search_box.send_keys(parcel)
            driver.find_element_by_xpath("//a[@id='ctlBodyPane_ctl04_ctl01_btnSearch']").click()
            # Summary - Auditor's Office
            # Click random Item
            if randomoption == 1:
                collapse = collapseID1 + str(randomCnum )+ collapseID2
                if len(driver.find_elements_by_xpath("//span[@id = '" + collapse + "']")):
                    element = driver.find_element_by_xpath("//span[@id = '" + collapse + "']")
                    driver.execute_script("return arguments[0].scrollIntoView();", element)    
                    element.click()
                    time.sleep(0.5)                
                    element.click()
                    time.sleep(randomTime)
            if len(driver.find_elements_by_xpath("//section[@id = 'ctlBodyPane_ctl00_mSection']")):
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl00_ctl01_lblAlternateID']")):
                    valueList[0] = driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl00_ctl01_lblAlternateID']").text
                    # worksheet.cell(row = row_count, column = 49).value =  altKey
                    
                    # print('altKey--  ', altKey)
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl00_ctl01_lblPropertyAddress']")):
                    valueList[1] = driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl00_ctl01_lblPropertyAddress']").text
                    valueList[1] = valueList[1].replace('\n', ', ')
                    # worksheet.cell(row = row_count, column = 50).value = propertyAddress
                    # print('property--  ', propertyAddress)
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl00_ctl01_lblLegalDescription']")):
                    valueList[2] = driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl00_ctl01_lblLegalDescription']").text
                    # worksheet.cell(row = row_count, column = 51).value = briefDescription
                    # print('Description--  ', briefDescription)
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl00_ctl01_lblClass']")):
                    valueList[3] = driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl00_ctl01_lblClass']").text
                    # worksheet.cell(row = row_count, column = 52).value = classItem
                    # print('class--  ', classItem)
            # Owners - Auditor's Office
            ownerAddress = []
            owner = ''
            # Click random Item
            if randomoption == 2:
                collapse = collapseID1 + str(randomCnum )+ collapseID2
                if len(driver.find_elements_by_xpath("//span[@id = '" + collapse + "']")):
                    element = driver.find_element_by_xpath("//span[@id = '" + collapse + "']")
                    driver.execute_script("return arguments[0].scrollIntoView();", element)                    
                    element.click()
                    time.sleep(0.5)
                    element.click()
                    time.sleep(randomTime)
            if len(driver.find_elements_by_xpath("//section[@id = 'ctlBodyPane_ctl02_mSection']")):
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lblDeedName_lblSearch']")):
                    valueList[4] = driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lblDeedName_lblSearch']").text
                    # worksheet.cell(row = row_count, column = 53).value = ownerName
                    # print('ownername--  ', valueList[4])
                if len(driver.find_elements_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lblDeedName_lnkSearch']")):
                    valueList[4] = driver.find_element_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lblDeedName_lnkSearch']").text
                    # worksheet.cell(row = row_count, column = 53).value = ownerName
                    # print('ownername--  ', ownerName)
                if len(driver.find_elements_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lnkAddress1']")):
                    valueList[5] = driver.find_element_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lnkAddress1']").text + ', '
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lblAddress1']")):
                    valueList[5] = driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lblAddress1']").text + ', '  
                
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lblAddress2']")):
                    valueList[5] += driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lblAddress2']").text + ', '
                
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lblAddress3']")):
                    valueList[5] += driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl01_lblAddress3']").text + ', '
                valueList[5] = valueList[5][:-2]
                print('ownername--  ', valueList[4])
                print('owneraddr--  ', valueList[5])
                # worksheet.cell(row = row_count, column = 54).value = owner[:-2]

                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lblDeedName_lblSearch']")):
                    valueList[6] = driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lblDeedName_lblSearch']").text
                    # worksheet.cell(row = row_count, column = 53).value = ownerName
                    # print('ownername--  ', ownerName)
                if len(driver.find_elements_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lblDeedName_lnkSearch']")):
                    valueList[6] = driver.find_element_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lblDeedName_lnkSearch']").text
                    # worksheet.cell(row = row_count, column = 53).value = ownerName
                    # print('ownername--  ', ownerName)
                
                if len(driver.find_elements_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lnkAddress1']")):
                    valueList[7] = driver.find_element_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lnkAddress1']").text + ', '
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lblAddress1']")):
                    valueList[7] += driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lblAddress1']").text + ', '  
                
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lblAddress2']")):
                    valueList[7] += driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lblAddress2']").text + ', '
                
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lblAddress3']")):
                    valueList[7] += driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl03_lblAddress3']").text + ', '
                valueList[7] = valueList[7][:-2]
                # worksheet.cell(row = row_count, column = 54).value = owner[:-2]

                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lblDeedName_lblSearch']")):
                    valueList[8] = driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lblDeedName_lblSearch']").text
                    # worksheet.cell(row = row_count, column = 53).value = ownerName
                    # print('ownername--  ', ownerName)
                if len(driver.find_elements_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lblDeedName_lnkSearch']")):
                    valueList[8] = driver.find_element_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lblDeedName_lnkSearch']").text
                    # worksheet.cell(row = row_count, column = 53).value = ownerName
                    # print('ownername--  ', ownerName)
                
                if len(driver.find_elements_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lnkAddress1']")):
                    valueList[9] = driver.find_element_by_xpath("//a[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lnkAddress1']").text + ', '
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lblAddress1']")):
                    valueList[9] += driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lblAddress1']").text + ', '  
                
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lblAddress2']")):
                    valueList[9] += driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lblAddress2']").text + ', '
                
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lblAddress3']")):
                    valueList[9] += driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl02_ctl01_lstDeed_ctl05_lblAddress3']").text + ', '
                valueList[9] = valueList[9][:-2]
                # worksheet.cell(row = row_count, column = 54).value = owner[:-2]

            # Residential Dwellings - Assessor's Office
            # Click random Item
            if randomoption == 3:
                collapse = collapseID1 + str(randomCnum )+ collapseID2
                if len(driver.find_elements_by_xpath("//span[@id = '" + collapse + "']")):
                    element = driver.find_element_by_xpath("//span[@id = '" + collapse + "']")
                    driver.execute_script("return arguments[0].scrollIntoView();", element)  
                    element.click()                  
                    time.sleep(0.5)
                    element.click()
                    time.sleep(randomTime)
            if len(driver.find_elements_by_xpath("//section[@id = 'ctlBodyPane_ctl06_mSection']")):
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl06_ctl01_lstResidential_ctl00_lblOccupancy']")):
                    valueList[10] = driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl06_ctl01_lstResidential_ctl00_lblOccupancy']").text
                    # worksheet.cell(row = row_count, column = 55).value = occupancy
                    # print('Occupancy--  ', occupancy)
                if len(driver.find_elements_by_xpath("//span[@id = 'ctlBodyPane_ctl06_ctl01_lstResidential_ctl00_lblYearBuilt']")):
                    valueList[11] = driver.find_element_by_xpath("//span[@id = 'ctlBodyPane_ctl06_ctl01_lstResidential_ctl00_lblYearBuilt']").text
                    # worksheet.cell(row = row_count, column = 56).value = yearBuilt
                    # print('Year Built--  ', yearBuilt)
            # Sales - Assessor's Office
            # Click random Item
            if randomoption == 4:
                collapse = collapseID1 + str(randomCnum )+ collapseID2
                if len(driver.find_elements_by_xpath("//span[@id = '" + collapse + "']")):
                    element = driver.find_element_by_xpath("//span[@id = '" + collapse + "']")
                    driver.execute_script("return arguments[0].scrollIntoView();", element)                    
                    element.click()
                    time.sleep(0.5)
                    element.click()
                    time.sleep(randomTime)
            if len(driver.find_elements_by_xpath("//section[@id = 'ctlBodyPane_ctl12_mSection']")):
                if len(driver.find_elements_by_xpath("//table[@id = 'ctlBodyPane_ctl12_ctl01_gvwSales']/tbody/tr[1]")):
                    valueList[12] = driver.find_element_by_xpath("//table[@id = 'ctlBodyPane_ctl12_ctl01_gvwSales']/tbody/tr[1]/td[1]").text
                    valueList[13] = driver.find_element_by_xpath("//table[@id = 'ctlBodyPane_ctl12_ctl01_gvwSales']/tbody/tr[1]/td[8]").text
                    # worksheet.cell(row = row_count, column = 57).value = saleDate
                    # worksheet.cell(row = row_count, column = 58).value = amount
                    # print('saleDate--  ', saleDate)
                    # print('Amount--  ', amount)
            # Valuation - Assessor's Office
            # Click random Item
            if randomoption == 5:
                collapse = collapseID1 + str(randomCnum )+ collapseID2
                if len(driver.find_elements_by_xpath("//span[@id = '" + collapse + "']")):
                    element = driver.find_element_by_xpath("//span[@id = '" + collapse + "']")
                    driver.execute_script("return arguments[0].scrollIntoView();", element) 
                    element.click()   
                    time.sleep(1)                
                    element.click()
                    time.sleep(randomTime) 
            if len(driver.find_elements_by_xpath("//section[@id = 'ctlBodyPane_ctl14_mSection']")):
                if len(driver.find_elements_by_xpath("//section[@id = 'ctlBodyPane_ctl14_mSection']/div/table")):
                    valueList[14] = driver.find_element_by_xpath("//section[@id = 'ctlBodyPane_ctl14_mSection']/div/table/tbody/tr[2]/td[3]").text
                    valueList[15] = driver.find_element_by_xpath("//section[@id = 'ctlBodyPane_ctl14_mSection']/div/table/tbody/tr[4]/td[3]").text
                    valueList[16] = driver.find_element_by_xpath("//section[@id = 'ctlBodyPane_ctl14_mSection']/div/table/tbody/tr[7]/td[3]").text
                    # worksheet.cell(row = row_count, column = 59).value = landValue
                    # worksheet.cell(row = row_count, column = 60).value = BLDGValue
                    # worksheet.cell(row = row_count, column = 61).value = justValue
                    # print('landValue--  ', landValue)
                    # print('BLDGValue--  ', BLDGValue)
                    # print('justValue--  ', justValue)
                else:
                    print('no Section14')
            for idx in range(0, len(valueList)):
                worksheet.cell(row = row_count, column = idx + 49).value = valueList[idx]
            print("<><><><><><><><><><><><><><><><><><><><><><><><>")
        workbook.save('SCOTT COUNTY IA1.xlsx')
        driver.quit()